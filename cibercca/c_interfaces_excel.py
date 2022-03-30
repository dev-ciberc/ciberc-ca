from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ExcelInterfaces():

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.create_sheet("interfaces", 0)
        self.font_header = Font(color='ffffff', bold=True)

    def setHeadCol(self, ws, col, name):
        self.ws[col] = name
        self.ws[col].fill = PatternFill(
            start_color='000000', end_color='000000', fill_type="solid")
        self.ws[col].font = self.font_header
        return ws

    def setHeadColInterface(self, ws, col, name):
        self.ws[col] = name
        self.ws[col].fill = PatternFill(
            start_color='eccd6a', end_color='eccd6a', fill_type="solid")
        self.ws[col].font = self.font_header
        return ws

    def getvalueFroomDictInterface(self, data, device, interface, name):
        value = None
        try:
            value = data[device][interface][name]
        except Exception:
            pass

        return value

    def getValueFromArrayService(self, service, name):
        valor = None
        try:
            for item in service['data']:
                try:
                    valor = (item[name])
                except Exception:
                    pass
        except Exception:
            pass

        return valor

    def create(self, data, mechanism: str = "row"):

        self.setHeadCol(self.ws, 'A1',  "Interfaz antigua")
        self.setHeadCol(self.ws, 'B1', "Interfaz nueva")
        self.setHeadCol(self.ws, 'C1', "Negotation / STP")
        self.setHeadCol(self.ws, 'D1', "BVI")
        self.setHeadCol(self.ws, 'E1', "BD")  # bridge_domain
        self.setHeadCol(self.ws, 'F1', "DOT1Q")
        self.setHeadCol(self.ws, 'G1', "DESCRIPTION")
        self.setHeadCol(self.ws, 'H1', "MTU")
        self.setHeadCol(self.ws, 'I1', "VRF")
        self.setHeadCol(self.ws, 'J1', "XCONNECTION")
        self.setHeadCol(self.ws, 'K1', "XCONNECTION VC ID")
        self.setHeadCol(self.ws, 'L1', "PW-CLASS")
        self.setHeadCol(self.ws, 'M1', "QoS INPUT")
        self.setHeadCol(self.ws, 'N1', "QoS OUTPUT")
        self.setHeadCol(self.ws, 'O1', "IP ADDRESS")

        if mechanism != "column":
            # primary, secundary
            self.setHeadCol(self.ws, 'P1', "TYPE IP ADDRESS")

        # incicio de fila en el excel
        index = 2

        # --
        # -- PRIMERA iteracion son los resultados en los device (hosts)
        # --
        for device in data:

            EC_NAME_DEVICE = device  # noqa

            # --
            # -- SEGUNDA iteracion con las interfaces
            # --
            for interface in data[device]:

                EC_INTERFACE_NAME = interface
                EC_INTERFACE_DESCRIPTION = self.getvalueFroomDictInterface(data, device, interface, 'description')  # noqa
                EC_INTERFACE_MTU = self.getvalueFroomDictInterface(data, device, interface, 'mtu')  # noqa

                if EC_INTERFACE_NAME == "error":
                    break

                # primera iteracion con excel va solo la interface
                self.ws[f"A{index}"] = EC_INTERFACE_NAME
                self.ws[f"G{index}"] = EC_INTERFACE_DESCRIPTION
                self.ws[f"H{index}"] = EC_INTERFACE_MTU
                index += 1

                # --
                # -- TERCER iteracion, recorre los servicios por interface
                # --
                try:
                    services_list = [data[device][interface]["services"]] if data[device][interface]["services"].__class__ == dict else data[device][interface]["services"]  # noqa
                except Exception:
                    services_list = []

                if services_list != [{}]:  # noqa
                    for service in services_list:
                        EC_SERVICE_BRIDGE_DOMANIN = self.getValueFromArrayService(service, 'bridge_domain')  # noqa
                        EC_SERVICE_DOT1Q = self.getValueFromArrayService(service, 'dot1q')  # noqa
                        EC_SERVICE_QOS_INPUT = self.getValueFromArrayService(service, 'service_policy_input')  # noqa
                        EC_SERVICE_QOS_OUTPUT = self.getValueFromArrayService(service, 'service_policy_output')  # noqa

                        EC_SERVICE_XCONNECTIONS = self.getValueFromArrayService(service, 'xconnect_ip')  # noqa
                        EC_SERVICE_XCONNECTIONS_VC_ID = self.getValueFromArrayService(service, 'xconnect_id')  # noqa
                        EC_SERVICE_PW_CLASS = self.getValueFromArrayService(service, 'xconnect_class')  # noqa
                        EC_SERVICE_PW_CLASS_DATA = self.getValueFromArrayService(service, 'xconnect_class_data')  # noqa

                        EC_SERVICE_DESCRIPTION = self.getValueFromArrayService(service, "description")  # noqa
                        EC_SERVICE_MODE_TRUNK = self.getValueFromArrayService(service, "switchport_mode")  # noqa

                        # --
                        # -- CUARTA iteracion, recorre los servicios por interface # noqa: E501
                        # --

                        try:
                            vlans_list = [service['vlan']] if service['vlan'].__class__ == dict else service['vlan']  # noqa
                        except Exception:
                            vlans_list = []

                        if len(vlans_list) > 0 and vlans_list != [{}]:
                            for vlan in vlans_list:
                                # NOTA:
                                # service_policy_input, service_policy_output (son del servicio) # noqa
                                # si se encuentra en la vlan se omite, debe ser del servicio # noqa
                                # ip_helper_address => va como nota en el excel

                                try:
                                    EC_VLAN_DESCRIPTION = vlan['description']['description']  # noqa
                                except Exception:
                                    EC_VLAN_DESCRIPTION = None

                                try:
                                    EC_VLAN_VRF = vlan['vrf']['vrf']  # noqa
                                except Exception:
                                    EC_VLAN_VRF = None  # noqa

                                try:
                                    EC_VLAN_QOS_INPUT = vlan['service_policy_input']['service_policy_input']  # noqa
                                except Exception:
                                    EC_VLAN_QOS_INPUT = None  # noqa

                                try:
                                    EC_VLAN_QOS_OUTPUT = vlan['service_policy_output']['service_policy_output']  # noqa
                                except Exception:
                                    EC_VLAN_QOS_OUTPUT = None  # noqa

                                try:
                                    EC_VLAN_IP_ADDRESS_PRIMARY = vlan['ip_address_primary']['ip_address']  # noqa
                                    EC_VLAN_IP_ADDRESS_MASK = vlan['ip_address_primary']['mask']  # noqa

                                    if EC_VLAN_IP_ADDRESS_PRIMARY is not None:
                                        EC_VLAN_PRIMARY_IP = f"{EC_VLAN_IP_ADDRESS_PRIMARY} {EC_VLAN_IP_ADDRESS_MASK}"  # noqa
                                        EC_VLAN_IP_ADDRESS_TYPE = 'primary'
                                except Exception:
                                    EC_VLAN_PRIMARY_IP = None
                                    EC_VLAN_IP_ADDRESS_TYPE = None

                                # --
                                # -- [VALIDACION] - saber si el servicio es BVI
                                EC_BVI = None
                                IS_BVI = "BVI"
                                try:
                                    # los troncales no pueden ser BVI
                                    if EC_SERVICE_MODE_TRUNK is not None:
                                        try:
                                            cant_ports = vlan['cant_ports']
                                            cant_bridges = vlan['cant_bridges']
                                        except Exception:
                                            cant_ports = 0
                                            cant_bridges = 0

                                        # es BVI si tiene mas de dos puertos, con mas de dos bridges # noqa
                                        if cant_ports > 0 and cant_bridges > 1:
                                            EC_BVI = IS_BVI

                                        # si el servicio es de gestion debe tener mas de dos bridges # noqa
                                        # para poder ser BVI
                                        elif EC_VLAN_DESCRIPTION.lower().find("gest") and cant_bridges > 1:  # noqa
                                            EC_BVI = IS_BVI

                                        # sera BVI si es capa2 con mas de dos bridges # noqa
                                        elif EC_SERVICE_XCONNECTIONS is not None and cant_bridges > 1:  # noqa
                                            EC_BVI = IS_BVI

                                        # si todas las validaciones no se encuentran y existen # noqa
                                        # mas de dos puertos en la mac_address table es BVI # noqa
                                        elif cant_ports > 1:
                                            EC_BVI = IS_BVI

                                    # --
                                    # --
                                    # se aprovecha la verificacion de troncal
                                    else:
                                        # --
                                        # solo para vlan que se comportan como troncales  # noqa
                                        # si es modo trocal, el xconnect se verifica en vlan  # noqa
                                        try:
                                            xconnect = vlan['xconnect']
                                            EC_SERVICE_XCONNECTIONS = xconnect['xconnect_ip']  # noqa
                                            EC_SERVICE_XCONNECTIONS_VC_ID = xconnect['xconnect_id']  # noqa
                                            EC_SERVICE_PW_CLASS = xconnect['xconnect_class']  # noqa
                                            EC_SERVICE_PW_CLASS_DATA = xconnect['xconnect_class_data']  # noqa
                                        except Exception:
                                            pass
                                except Exception:
                                    pass

                                self.ws[f"A{index}"] = EC_INTERFACE_NAME
                                self.ws[f"D{index}"] = EC_BVI
                                self.ws[f"E{index}"] = EC_SERVICE_BRIDGE_DOMANIN  # noqa
                                self.ws[f"F{index}"] = EC_SERVICE_DOT1Q
                                self.ws[f"G{index}"] = EC_VLAN_DESCRIPTION

                                self.ws[f"I{index}"] = EC_VLAN_VRF
                                self.ws[f"J{index}"] = EC_SERVICE_XCONNECTIONS
                                self.ws[f"K{index}"] = EC_SERVICE_XCONNECTIONS_VC_ID  # noqa
                                self.ws[f"L{index}"] = EC_SERVICE_PW_CLASS_DATA if EC_SERVICE_PW_CLASS == 'pw_class' else None  # noqa

                                self.ws[f"M{index}"] = EC_SERVICE_QOS_INPUT  # if EC_VLAN_QOS_INPUT == None else EC_VLAN_QOS_INPUT  # noqa
                                self.ws[f"N{index}"] = EC_SERVICE_QOS_OUTPUT  # if EC_VLAN_QOS_OUTPUT == None else EC_VLAN_QOS_OUTPUT  # noqa
                                self.ws[f"O{index}"] = EC_VLAN_PRIMARY_IP

                                if mechanism == 'row':
                                    self.ws[f"P{index}"] = EC_VLAN_IP_ADDRESS_TYPE  # noqa
                                    index += 1

                                    try:
                                        ip_secundary_list = [vlan['ip_address_secundary']] if vlan['ip_address_secundary'].__class__ == dict else vlan['ip_address_secundary']  # noqa
                                    except Exception:
                                        ip_secundary_list = []

                                    if ip_secundary_list != [{}]:
                                        for secundary in ip_secundary_list:

                                            try:
                                                EC_VLAN_IP_ADDRESS = secundary['ip_address']  # noqa
                                                EC_VLAN_IP_ADDRESS_MASK = secundary['mask']  # noqa
                                                EC_VLAN_IP_ADDRESS_TYPE = secundary['type']  # noqa

                                                if EC_VLAN_IP_ADDRESS is not None:  # noqa
                                                    EC_VLAN_IP_SECUNDARY = f"{EC_VLAN_IP_ADDRESS} {EC_VLAN_IP_ADDRESS_MASK}"  # noqa
                                            except Exception:
                                                EC_VLAN_IP_SECUNDARY = None

                                            self.ws[f"A{index}"] = EC_INTERFACE_NAME  # noqa
                                            self.ws[f"D{index}"] = EC_BVI  # noqa
                                            self.ws[f"E{index}"] = EC_SERVICE_BRIDGE_DOMANIN  # noqa
                                            self.ws[f"F{index}"] = EC_SERVICE_DOT1Q  # noqa
                                            self.ws[f"G{index}"] = EC_VLAN_DESCRIPTION  # noqa

                                            self.ws[f"I{index}"] = EC_VLAN_VRF  # noqa
                                            self.ws[f"J{index}"] = EC_SERVICE_XCONNECTIONS  # noqa
                                            self.ws[f"K{index}"] = EC_SERVICE_XCONNECTIONS_VC_ID  # noqa
                                            self.ws[f"L{index}"] = EC_SERVICE_PW_CLASS_DATA if EC_SERVICE_PW_CLASS == 'pw_class' else None  # noqa

                                            self.ws[f"M{index}"] = EC_SERVICE_QOS_INPUT  # if EC_VLAN_QOS_INPUT == None else EC_VLAN_QOS_INPUT  # noqa
                                            self.ws[f"N{index}"] = EC_SERVICE_QOS_OUTPUT  # if EC_VLAN_QOS_OUTPUT == None else EC_VLAN_QOS_OUTPUT  # noqa
                                            self.ws[f"O{index}"] = EC_VLAN_IP_SECUNDARY  # noqa
                                            self.ws[f"P{index}"] = EC_VLAN_IP_ADDRESS_TYPE  # noqa
                                            index += 1

                                else:

                                    try:
                                        ip_secundary_list = [vlan['ip_address_secundary']] if vlan['ip_address_secundary'].__class__ == dict else vlan['ip_address_secundary']  # noqa
                                    except Exception:
                                        ip_secundary_list = []

                                    if ip_secundary_list != [{}]:

                                        COLUMN = 16
                                        NUM_IP = 2
                                        for secundary in ip_secundary_list:

                                            try:
                                                EC_VLAN_IP_ADDRESS = secundary['ip_address']  # noqa
                                                EC_VLAN_IP_ADDRESS_MASK = secundary['mask']  # noqa
                                                EC_VLAN_IP_ADDRESS_TYPE = secundary['type']  # noqa

                                                if EC_VLAN_IP_ADDRESS is not None:  # noqa
                                                    EC_VLAN_IP_SECUNDARY = f"{EC_VLAN_IP_ADDRESS} {EC_VLAN_IP_ADDRESS_MASK} {EC_VLAN_IP_ADDRESS_TYPE}"  # noqa
                                            except Exception:
                                                EC_VLAN_IP_SECUNDARY = None

                                            self.ws.cell(row=1, column=COLUMN).value = f'IP ADDRESS {NUM_IP}'  # noqa
                                            self.ws.cell(row=index, column=COLUMN).value = EC_VLAN_IP_SECUNDARY  # noqa
                                            COLUMN += 1
                                            NUM_IP += 1

                                    # fin por multiples columnas en excel
                                    index += 1

                        else:
                            # --
                            # -- si el servicio no tiene vlans programdas
                            # datos servicios
                            self.ws[f"A{index}"] = EC_INTERFACE_NAME
                            self.ws[f"E{index}"] = EC_SERVICE_BRIDGE_DOMANIN
                            self.ws[f"F{index}"] = EC_SERVICE_DOT1Q
                            self.ws[f"G{index}"] = EC_SERVICE_DESCRIPTION

                            # --
                            # -- [validacion] - se pega el dato solo si es pw_class # noqa
                            # --
                            self.ws[f"J{index}"] = EC_SERVICE_XCONNECTIONS
                            self.ws[f"K{index}"] = EC_SERVICE_XCONNECTIONS_VC_ID  # noqa
                            self.ws[f"L{index}"] = EC_SERVICE_PW_CLASS_DATA if EC_SERVICE_PW_CLASS == 'pw_class' else None  # noqa

                            self.ws[f"M{index}"] = EC_SERVICE_QOS_INPUT
                            self.ws[f"N{index}"] = EC_SERVICE_QOS_OUTPUT
                            index += 1

    def save(self, name):
        self.wb.save(f"{name}.xlsx")
