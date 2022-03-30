from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ExcelVrfDestination():

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.create_sheet("interfaces", 0)
        self.font_header = Font(color='ffffff', bold=True)
        self.index = 2
        self.na = "N/A"

    def setHeadCol(self, ws, col, name):
        self.ws[col] = name
        self.ws[col].fill = PatternFill(
            start_color='000000', end_color='000000', fill_type="solid")
        self.ws[col].font = self.font_header
        return ws

    def create(self, data):

        self.setHeadCol(self.ws, 'A1', "device")
        self.setHeadCol(self.ws, 'B1', "hostname")
        self.setHeadCol(self.ws, 'C1', "vrf")
        self.setHeadCol(self.ws, 'D1', "learning interface")
        self.setHeadCol(self.ws, 'E1', "neighbor mac address")
        self.setHeadCol(self.ws, 'F1', "arp neighbor")
        self.setHeadCol(self.ws, 'G1', "ping result")

        for key in data:
            device = data[key]['data']

            EC_DEVICE_NAME = device['name']
            EC_DEVICE_HOSTNAME = device['hostname']

            # --
            # listar las vrfs asignadas al equipo origen
            # --

            for vrf in device['list_vrf']:
                EC_VRF_NAME = vrf['vrf_name']

                if len(vrf['ips_arp']) > 0:
                    for ip in vrf["ips_arp"]:
                        EC_VRF_INTERFACE = ip['interface']
                        EC_VRF_MAC_ADDRESS = ip['mac']
                        EC_VRF_ARP_NEIGBORD = ip['ip_address']
                        EC_VRF_PING_PERCENT = ip["ping_arp"][0]['percent']

                        self.ws['A' + str(self.index)] = EC_DEVICE_NAME
                        self.ws['B' + str(self.index)] = EC_DEVICE_HOSTNAME
                        self.ws['C' + str(self.index)] = EC_VRF_NAME
                        self.ws['D' + str(self.index)] = EC_VRF_INTERFACE
                        self.ws['E' + str(self.index)] = EC_VRF_MAC_ADDRESS
                        self.ws['F' + str(self.index)] = EC_VRF_ARP_NEIGBORD
                        self.ws['G' + str(self.index)] = EC_VRF_PING_PERCENT

                        self.index += 1
                else:
                    self.ws['A' + str(self.index)] = EC_DEVICE_NAME
                    self.ws['B' + str(self.index)] = EC_DEVICE_HOSTNAME
                    self.ws['C' + str(self.index)] = EC_VRF_NAME
                    self.ws['D' + str(self.index)] = self.na
                    self.ws['E' + str(self.index)] = self.na
                    self.ws['F' + str(self.index)] = self.na
                    self.ws['G' + str(self.index)] = self.na

                    self.index += 1

    def save(self, name):
        self.wb.save(f"{name}.xlsx")
